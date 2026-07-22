"""
sales_growth/build_excel.py
Emit the consolidated Sales Growth plan as an .xlsx that mirrors the Strategic
Initiatives List columns. Two tabs:
  - "List Import" : exact List columns, Initiative rows + Task rows (a Parent Ref
    helper column names each task's parent; the live Parent ID is set by the seed
    script / dashboard, since SharePoint assigns the real numeric ID on create).
  - "Plan Overview" : a readable view with the key decisions per initiative.
Writes to the user's Downloads folder.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from plan_data import PLAN, PLAN_START

# List columns (friendly headers), in List order. ID / Created By / Modified By are
# system-assigned on create, so they're omitted from an import sheet.
LIST_COLS = [
    "Task/Initiative","Sponsor","Owner","Region","Category","Priority","Initiative Name",
    "Description","Status","% Complete","Completed Actions","Next Action","Blockers",
    "Start Date","Target Completion","Revised Completion","Actual Completion",
    "Forecasted Cost","Realized Cost","Forecasted Revenue Impact","Realized Revenue To Date",
    "Forecasted EBITDA Impact","Realized EBITDA To Date","Benefit Start Date","Project Link",
    "Approval status",
]
HELPER_COLS = ["Ref","Parent Ref"]   # not List columns; drop before an actual import

BLUE="0067B1"; DARK="1F3A4D"; CORE="0067B1"; EXPAND="B8791F"; ENABLE="2E7B87"
GREY="F1F5F9"; LINE="D0D7DE"
TIER_COLOR={"Core Play":CORE,"Expansion Bet":EXPAND,"Enabler":ENABLE}


def _desc(i):
    return (f"{i['desc']}  |  KPI: {i['kpi']}  |  Contributors: {', '.join(i['contributors'])}")


def _rows():
    """Yield dicts keyed by header for every Initiative and Task row."""
    for i in PLAN:
        yield {
            "Task/Initiative":"Initiative","Sponsor":i["sponsor"],"Owner":i["owner"],
            "Region":i["region"],"Category":"BOD","Priority":i["priority"],
            "Initiative Name":i["name"],"Description":_desc(i),"Status":"Not Started",
            "% Complete":0,"Next Action":"DECISION: "+i["decisions"][0],
            "Start Date":PLAN_START,"Target Completion":i["target"],
            "Forecasted Revenue Impact":i["rev"] or "",
            "Forecasted EBITDA Impact":i["ebitda"] or "",
            "Ref":i["code"],"Parent Ref":"",
        }
        for name, phase, due in i["tasks"]:
            label = "90-day action" if phase=="90d" else "Year-one milestone"
            yield {
                "Task/Initiative":"Task","Sponsor":i["sponsor"],"Owner":i["owner"],
                "Region":i["region"],"Category":"BOD","Priority":i["priority"],
                "Initiative Name":name,"Description":label,"Status":"Not Started",
                "% Complete":0,"Start Date":PLAN_START,"Target Completion":due,
                "Ref":"","Parent Ref":f"{i['code']} - {i['name']}",
            }


def build(path):
    wb=openpyxl.Workbook()
    thin=Side(style="thin",color=LINE); bd=Border(thin,thin,thin,thin)
    cols=LIST_COLS+HELPER_COLS

    # ---- Tab 1: List Import ----
    ws=wb.active; ws.title="List Import"
    ws["A1"]="IEG Sales Growth - Strategic Initiatives (List import)"
    ws["A1"].font=Font(name="Segoe UI",size=13,bold=True,color="FFFFFF")
    ws["A1"].fill=PatternFill("solid",fgColor=BLUE); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(cols))
    ws.row_dimensions[1].height=22
    ws["A2"]=("Category=BOD, Region=Corp (UK for EMEA). Initiative rows first, then their Task rows. "
              "'Parent Ref' names each task's parent; the live Parent ID is set on create. Drop the two helper columns before importing.")
    ws["A2"].font=Font(name="Segoe UI",italic=True,size=9,color="555555"); ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=len(cols))
    hdr=3
    for j,h in enumerate(cols,1):
        c=ws.cell(row=hdr,column=j,value=h)
        c.font=Font(bold=True,color="FFFFFF",size=9)
        c.fill=PatternFill("solid",fgColor=(DARK if h not in HELPER_COLS else "6B7B8A"))
        c.alignment=Alignment(horizontal="center",wrap_text=True)
    r=hdr+1
    for row in _rows():
        is_ini=row["Task/Initiative"]=="Initiative"
        for j,h in enumerate(cols,1):
            v=row.get(h,"")
            c=ws.cell(row=r,column=j,value=v); c.border=bd
            c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=(h in ("Initiative Name","Description","Next Action","Parent Ref")))
            if h in ("Forecasted Revenue Impact","Forecasted EBITDA Impact") and isinstance(v,(int,float)) and v!="":
                c.number_format='$#,##0'
        if is_ini:
            tier=next(x["tier"] for x in PLAN if x["code"]==row["Ref"])
            fill=PatternFill("solid",fgColor=GREY)
            for j in range(1,len(cols)+1):
                ws.cell(row=r,column=j).fill=fill
                ws.cell(row=r,column=j).font=Font(bold=True,size=10)
            ws.cell(row=r,column=1).font=Font(bold=True,color=TIER_COLOR[tier])
        r+=1
    widths=[13,15,15,9,9,7,34,52,11,9,16,30,12,11,13,12,12,12,12,15,15,15,15,12,11,12,7,26]
    for j,w in enumerate(widths[:len(cols)],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
    ws.freeze_panes="A4"

    # ---- Tab 2: Plan Overview (readable, with decisions) ----
    ws2=wb.create_sheet("Plan Overview")
    ws2["A1"]="IEG Sales Growth - Consolidated Plan (overview)"
    ws2["A1"].font=Font(name="Segoe UI",size=13,bold=True,color="FFFFFF")
    ws2["A1"].fill=PatternFill("solid",fgColor=BLUE); ws2.merge_cells("A1:F1"); ws2.row_dimensions[1].height=22
    oc=["Ref","Initiative","Tier","Sponsor","Owner","Priority","Forecast Rev","Forecast EBITDA","KPI","Key decisions","90-day actions","Year-one milestones"]
    for j,h in enumerate(oc,1):
        c=ws2.cell(row=3,column=j,value=h); c.font=Font(bold=True,color="FFFFFF",size=9)
        c.fill=PatternFill("solid",fgColor=DARK); c.alignment=Alignment(horizontal="center",wrap_text=True)
    r=4
    for i in PLAN:
        d90=[f"- {n}" for n,p,_ in i["tasks"] if p=="90d"]
        dyr=[f"- {n}" for n,p,_ in i["tasks"] if p=="year"]
        dec=[f"- {x}" for x in i["decisions"]]
        vals=[i["code"],i["name"],i["tier"],i["sponsor"],i["owner"],i["priority"],
              i["rev"] or "", i["ebitda"] or "", i["kpi"], "\n".join(dec), "\n".join(d90), "\n".join(dyr)]
        for j,v in enumerate(vals,1):
            c=ws2.cell(row=r,column=j,value=v); c.border=bd
            c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
            if j in (7,8) and isinstance(v,(int,float)) and v!="": c.number_format='$#,##0'
        ws2.cell(row=r,column=1).font=Font(bold=True,color=TIER_COLOR[i["tier"]])
        ws2.cell(row=r,column=2).font=Font(bold=True)
        ws2.row_dimensions[r].height=max(72, 14*max(len(dec),len(d90),len(dyr)))
        r+=1
    for j,w in enumerate([6,30,13,15,15,7,13,14,34,46,52,46],1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(j)].width=w
    ws2.freeze_panes="A4"

    wb.properties.creator = "Chad Abrahamson"
    wb.properties.lastModifiedBy = "Chad Abrahamson"
    try: wb.save(path)
    except PermissionError:
        path=path.replace(".xlsx","_v2.xlsx"); wb.save(path)
    return path


if __name__ == "__main__":
    out=os.path.join(os.path.expanduser("~"),"Downloads","IEG_SLT_Initiatives_Import.xlsx")
    p=build(out)
    n_ini=len(PLAN); n_task=sum(len(i["tasks"]) for i in PLAN)
    print(f"saved {p}\n  {n_ini} initiatives + {n_task} tasks = {n_ini+n_task} rows | tabs: List Import, Plan Overview")
